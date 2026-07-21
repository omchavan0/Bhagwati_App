"""
============================================================================
DB STOCK IN — Reverse-GST Purchase Entry (MRP-based Auto Calculation)
============================================================================
db_purchase.py (Supplier-bill based) पेक्षा हे वेगळं आणि हलकं आहे — इथे
प्रत्येक स्टॉक-इन एन्ट्री MRP वरून मागच्या बाजूने (reverse) संपूर्ण गणित
आपोआप काढते (Base Rate, Buying Discount%, Taxable Value, CGST/SGST).
कुठलाही आकडा हार्डकोड नाही — सगळं Part Master मधल्या GST Rate वरून लाईव्ह.
============================================================================
"""
from datetime import datetime

from db_core import _get_connection, get_current_owner_uid, get_current_owner_email, get_device_id, _queue_sync


STOCK_IN_COLUMNS = [
    ("part_id", "INTEGER"),
    ("part_number", "TEXT"),
    ("description", "TEXT"),
    ("hsn_sac", "TEXT"),
    ("gst_rate", "REAL DEFAULT 18"),
    ("location", "TEXT"),
    ("mrp", "REAL DEFAULT 0"),
    ("qty", "REAL DEFAULT 0"),
    ("buy_rate", "REAL DEFAULT 0"),
    ("supplier_name", "TEXT"),
    ("base_rate", "REAL DEFAULT 0"),
    ("total_base", "REAL DEFAULT 0"),
    ("buying_percentage", "REAL DEFAULT 0"),
    ("taxable_value", "REAL DEFAULT 0"),
    ("cgst", "REAL DEFAULT 0"),
    ("sgst", "REAL DEFAULT 0"),
    ("total_amount", "REAL DEFAULT 0"),
    ("tx_date", "TEXT"),
    ("notes", "TEXT"),
    ("created_at", "TEXT"),
]


def _touch_sync_fields(record_id, table):
    conn = _get_connection()
    c = conn.cursor()
    c.execute(f"UPDATE {table} SET updated_at=?, device_id=?, owner_uid=?, owner_email=? WHERE id=?",
              (datetime.now().isoformat(), get_device_id(), get_current_owner_uid(),
               get_current_owner_email(), record_id))
    conn.commit()
    conn.close()


def init_table(conn, c):
    """stock_in_entries टेबल तयार/मायग्रेट करतं — db_core.init_db() कडून कॉल होतं."""
    c.execute('''CREATE TABLE IF NOT EXISTS stock_in_entries (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    part_id INTEGER
                 )''')
    conn.commit()

    c.execute("PRAGMA table_info(stock_in_entries)")
    existing = {row[1] for row in c.fetchall()}
    for col_name, col_type in STOCK_IN_COLUMNS:
        if col_name not in existing:
            c.execute(f"ALTER TABLE stock_in_entries ADD COLUMN {col_name} {col_type}")
    conn.commit()


# ======================================================================
# CORE FORMULA ENGINE — Government-Compliant Reverse GST Calculation
# ======================================================================

def calculate_stock_in(mrp, qty, buy_rate, gst_rate):
    """दिलेल्या MRP/Qty/Buy Rate/GST% वरून संपूर्ण Purchase गणित काढतो.
    कुठलाही आकडा हार्डकोड नाही — सगळं इथेच दिलेल्या पॅरामीटर्सवरून लाईव्ह
    काढलं जातं, त्यामुळे कुठलाही GST Rate (5%/12%/18%/28%) चालतो.

    सूत्र (Government-Compliant):
      1. Base Rate (Excl. GST) = MRP / (1 + GST%/100)
      2. Total Base Amount     = Base Rate * Qty
      3. Buying Discount %     = ((Base Rate - Buy Rate) / Base Rate) * 100
      4. Taxable Value         = Buy Rate * Qty   (हीच खरी खरेदी किंमत)
      5. CGST = SGST           = Taxable Value * (GST%/2) / 100
      6. Total Amount (Incl GST) = Taxable Value + CGST + SGST

    Returns: dict — सगळे calculated आकडे
    """
    mrp = float(mrp or 0)
    qty = float(qty or 0)
    buy_rate = float(buy_rate or 0)
    gst_rate = float(gst_rate or 0)

    base_rate = mrp / (1 + gst_rate / 100) if gst_rate else mrp
    total_base = base_rate * qty

    buying_percentage = ((base_rate - buy_rate) / base_rate * 100) if base_rate else 0.0

    taxable_value = buy_rate * qty
    half_gst = gst_rate / 2
    cgst = taxable_value * (half_gst / 100)
    sgst = taxable_value * (half_gst / 100)
    total_amount = taxable_value + cgst + sgst

    # Profit Margin — MRP आणि Buy Rate मधला फरक (ग्राहकाला विकतानाचा संभाव्य नफा)
    profit_per_unit = mrp - buy_rate
    profit_margin_percent = (profit_per_unit / mrp * 100) if mrp else 0.0

    return {
        "base_rate": base_rate,
        "total_base": total_base,
        "buying_percentage": buying_percentage,
        "taxable_value": taxable_value,
        "cgst": cgst,
        "sgst": sgst,
        "total_amount": total_amount,
        "profit_per_unit": profit_per_unit,
        "profit_margin_percent": profit_margin_percent,
    }


# ======================================================================
# SAVE — एक पूर्ण Stock-In एन्ट्री
# ======================================================================

def add_stock_in_entry(part_id, part_number, description, hsn_sac, gst_rate, location,
                        mrp, qty, buy_rate, supplier_name, tx_date="", notes=""):
    """एक Stock-In एन्ट्री सेव्ह करतो — गणित calculate_stock_in() मधून घेतो,
    Part Master चा buying_rate ताज्या दराने अपडेट करतो, आणि db_inventory.py
    च्या स्टॉक-लेजरमध्ये आपोआप stock-in जमा करतो (Module 1 शी सिंक)."""
    calc = calculate_stock_in(mrp, qty, buy_rate, gst_rate)

    conn = _get_connection()
    c = conn.cursor()
    c.execute('''INSERT INTO stock_in_entries
                 (part_id, part_number, description, hsn_sac, gst_rate, location,
                  mrp, qty, buy_rate, supplier_name, base_rate, total_base,
                  buying_percentage, taxable_value, cgst, sgst, total_amount,
                  tx_date, notes, created_at)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
              (part_id, part_number, description, hsn_sac, gst_rate, location,
               mrp, qty, buy_rate, supplier_name, calc["base_rate"], calc["total_base"],
               calc["buying_percentage"], calc["taxable_value"], calc["cgst"], calc["sgst"],
               calc["total_amount"], tx_date, notes, datetime.now().isoformat()))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    _touch_sync_fields(new_id, "stock_in_entries")
    _queue_sync("stock_in_entries", new_id, "upsert")

    # Module 1 शी सिंक — स्टॉक-लेजरमध्ये जमा + MRP/Buying Rate अपडेट
    # Module 1 शी सिंक — स्टॉक-लेजरमध्ये जमा + MRP/Buying Rate अपडेट
    if part_id:
        from db_inventory import record_stock_in, update_part_mrp_and_rate
        record_stock_in(part_id, qty, reference_table="stock_in_entries",
                         reference_id=new_id, tx_date=tx_date,
                         notes=f"Stock-In: {description} ({supplier_name})")
        update_part_mrp_and_rate(part_id, mrp, buy_rate)   # 👈 आता MRP सुद्धा जातं

    return new_id, calc


def get_stock_in_entries(limit=200):
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT * FROM stock_in_entries WHERE (is_deleted IS NULL OR is_deleted=0)
                 ORDER BY id DESC LIMIT ?''', (limit,))
    data = c.fetchall()
    conn.close()
    return data


def get_latest_stock_in_for_part(part_id):
    """त्या Part ची सगळ्यात अलीकडची Stock-In एन्ट्री — Profit Margin Sheet
    (Module 3) मध्ये सध्याचा MRP/Margin दाखवायला उपयोगी."""
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT * FROM stock_in_entries WHERE part_id=?
                 AND (is_deleted IS NULL OR is_deleted=0)
                 ORDER BY id DESC LIMIT 1''', (part_id,))
    row = c.fetchone()
    conn.close()
    return row

def get_inventory_sheet_rows():
    """Inventory Sheet साठी — Part Master + नवीनतम Stock-In एन्ट्री + Live
    Stock एकत्र. Stock-In नसेल (जुना Part) तर Part Master वर fallback."""
    from db_inventory import get_parts, get_part_stock

    rows = []
    for part in get_parts():
        stock = get_part_stock(part["id"])
        latest = get_latest_stock_in_for_part(part["id"])

        if latest:
            hsn_sac = latest["hsn_sac"] or ""
            gst_rate = latest["gst_rate"] if latest["gst_rate"] is not None else 18
            buy_rate = latest["buy_rate"] or 0
            mrp = latest["mrp"] or 0
            buy_dis_percent = latest["buying_percentage"] or 0
            vendor = latest["supplier_name"] or ""
            location = latest["location"] or ""
            last_stock_in_date = latest["tx_date"] or ""
        else:
            hsn_sac = part["hsn_sac"] if "hsn_sac" in part.keys() and part["hsn_sac"] else ""
            gst_rate = part["gst_rate"] if "gst_rate" in part.keys() and part["gst_rate"] is not None else 18
            buy_rate = part["buying_rate"] or 0
            mrp = part["mrp"] if "mrp" in part.keys() and part["mrp"] else 0
            buy_dis_percent = 0.0
            vendor = ""
            location = part["location"] if "location" in part.keys() and part["location"] else ""
            last_stock_in_date = ""

        reorder_level = part["reorder_level"] if "reorder_level" in part.keys() and part["reorder_level"] else 5
        brand = part["brand"] if "brand" in part.keys() and part["brand"] else ""
        category = part["category"] if "category" in part.keys() and part["category"] else ""
        unit = part["unit"] if "unit" in part.keys() and part["unit"] else "Nos"
        barcode = part["barcode"] if "barcode" in part.keys() and part["barcode"] else ""

        profit_per_unit = (mrp - buy_rate) if mrp else 0.0
        margin_percent = (profit_per_unit / mrp * 100) if mrp else 0.0
        total_amount = stock * buy_rate          # Stock Value — Buy Rate वर
        total_mrp_value = stock * mrp            # Stock Value — MRP वर (संभाव्य विक्री किंमत)

        is_out_of_stock = stock <= 0
        is_low_stock = (not is_out_of_stock) and reorder_level > 0 and stock <= reorder_level

        rows.append({
            "part_id": part["id"],
            "part_number": part["part_number"] or "",
            "description": part["product_name"],
            "hsn_sac": hsn_sac,
            "gst_rate": gst_rate,
            "buy_rate": buy_rate,
            "mrp": mrp,
            "qty": stock,
            "buy_dis_percent": buy_dis_percent,
            "total_amount": total_amount,
            "total_mrp_value": total_mrp_value,
            "profit_per_unit": profit_per_unit,
            "margin_percent": margin_percent,
            "vendor": vendor,
            "location": location,
            "brand": brand,
            "category": category,
            "unit": unit,
            "barcode": barcode,
            "reorder_level": reorder_level,
            "last_stock_in_date": last_stock_in_date,
            "is_out_of_stock": is_out_of_stock,
            "is_low_stock": is_low_stock,
        })

    rows.sort(key=lambda r: (r["description"] or "").lower())
    return rows


def get_inventory_sheet_totals(rows=None):
    rows = rows if rows is not None else get_inventory_sheet_rows()
    count = len(rows) or 1
    return {
        "total_qty": sum(r["qty"] for r in rows),
        "total_mrp": sum(r["mrp"] for r in rows),
        "total_buy_rate": sum(r["buy_rate"] for r in rows),
        "total_amount": sum(r["total_amount"] for r in rows),
        "total_mrp_value": sum(r["total_mrp_value"] for r in rows),
        "total_profit_per_unit": sum(r["profit_per_unit"] for r in rows),
        "avg_dis_percent": sum(r["buy_dis_percent"] for r in rows) / count,
        "avg_margin_percent": sum(r["margin_percent"] for r in rows) / count,
        "low_stock_count": sum(1 for r in rows if r["is_low_stock"]),
        "out_of_stock_count": sum(1 for r in rows if r["is_out_of_stock"]),
    }


def search_inventory_sheet_rows(query, filter_mode="all"):
    """filter_mode: 'all' | 'low_stock' | 'out_of_stock'"""
    rows = get_inventory_sheet_rows()
    if query and query.strip():
        q = query.strip().lower()
        rows = [r for r in rows if q in (r["part_number"] or "").lower() or q in (r["description"] or "").lower()]
    if filter_mode == "low_stock":
        rows = [r for r in rows if r["is_low_stock"]]
    elif filter_mode == "out_of_stock":
        rows = [r for r in rows if r["is_out_of_stock"]]
    return rows


def export_inventory_sheet_to_excel(filepath):
    """Inventory Sheet चा संपूर्ण Excel export — CA/Accountant किंवा स्टॉक
    ऑडिटसाठी उपयोगी."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = get_inventory_sheet_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Sheet"

    headers = ["Sr.No", "Part No", "Description", "HSN/SAC", "GST%", "MRP", "Buy Rate",
               "Qty", "Dis.%", "Amount", "Value@MRP", "Profit/Unit", "Margin%",
               "Vendor", "Location", "Brand", "Category", "Unit", "Barcode",
               "Reorder Level", "Last Stock-In"]
    ws.append(headers)
    header_fill = PatternFill("solid", start_color="00FFAA")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000"); cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, start=1):
        ws.append([
            i, r["part_number"], r["description"], r["hsn_sac"], r["gst_rate"], r["mrp"],
            r["buy_rate"], r["qty"], r["buy_dis_percent"], r["total_amount"], r["total_mrp_value"],
            r["profit_per_unit"], r["margin_percent"], r["vendor"], r["location"], r["brand"],
            r["category"], r["unit"], r["barcode"], r["reorder_level"], r["last_stock_in_date"],
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    wb.save(filepath)
    return filepath