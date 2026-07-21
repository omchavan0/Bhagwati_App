"""
============================================================================
DB UDHAARI — उधारी (Credit/Debit) रेकॉर्ड्स + रिपोर्ट्स
============================================================================
हे मॉड्यूल फक्त "udhaari" टेबल हाताळतं — add/update/delete, search, customer
history, due-date alerts, daily/monthly summary आणि Excel export. Isolation
मुळे यात बदल केला की expenses/daily_work/clients च्या कोडला काहीही धक्का
लागत नाही.
============================================================================
"""
from datetime import datetime, timedelta

from db_core import (
    _get_connection, _touch_sync_fields, _queue_sync, _parse_date,
    _safe_sync, _GSHEET_AVAILABLE, get_current_owner_uid, get_current_owner_email, get_device_id,
)

try:
    import gsheet_sync
except Exception:
    gsheet_sync = None

# udhaari टेबलचे सर्व कॉलम्स (id आणि name सोडून) — मायग्रेशन आणि इन्सर्टसाठी
COLUMNS = [
    ("mobile", "TEXT"),
    ("vehicle", "TEXT"),
    ("vehicle_no", "TEXT"),
    ("address", "TEXT"),
    ("tx_date", "TEXT"),
    ("due_date", "TEXT"),
    ("total_amt", "REAL DEFAULT 0"),
    ("paid_amt", "REAL DEFAULT 0"),
    ("due_amt", "REAL DEFAULT 0"),
    ("notes", "TEXT"),
    ("type", "TEXT DEFAULT 'Given'"),
    ("client_id", "INTEGER"),
    ("invoice_no", "TEXT"),              # 👈 नवीन
    ("customer_gstin", "TEXT"),          # 👈 नवीन
    ("customer_state_code", "TEXT"),     # 👈 नवीन
]

def _touch_sync_fields(record_id, table):
    conn = _get_connection()
    c = conn.cursor()
    c.execute(f"UPDATE {table} SET updated_at=?, device_id=?, owner_uid=?, owner_email=? WHERE id=?",
              (datetime.now().isoformat(), get_device_id(), get_current_owner_uid(),
               get_current_owner_email(), record_id))
    conn.commit()
    conn.close()

def init_table(conn, c):
    """udhaari टेबल तयार/मायग्रेट करतं — db_core.init_db() कडून कॉल होतं."""
    c.execute('''CREATE TABLE IF NOT EXISTS udhaari (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL
                 )''')
    conn.commit()

    c.execute("PRAGMA table_info(udhaari)")
    existing_cols = {row[1] for row in c.fetchall()}
    for col_name, col_type in COLUMNS:
        if col_name not in existing_cols:
            c.execute(f"ALTER TABLE udhaari ADD COLUMN {col_name} {col_type}")
    conn.commit()


# ======================================================================
# CRUD
# ======================================================================

def add_udhaari(name, mobile, vehicle, vehicle_no, address, tx_date,
                 due_date, total_amt, paid_amt, due_amt, notes, type, client_id=None,
                 invoice_no="", customer_gstin="", customer_state_code=""):
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''INSERT INTO udhaari
                 (name, mobile, vehicle, vehicle_no, address, tx_date,
                  due_date, total_amt, paid_amt, due_amt, notes, type, client_id,
                  invoice_no, customer_gstin, customer_state_code)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
              (name, mobile, vehicle, vehicle_no, address, tx_date,
               due_date, total_amt, paid_amt, due_amt, notes, type, client_id,
               invoice_no, customer_gstin, customer_state_code))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    _touch_sync_fields(new_id, "udhaari")
    _queue_sync("udhaari", new_id, "upsert")
    _safe_sync(gsheet_sync.sync_udhaari, get_udhaari()) if _GSHEET_AVAILABLE else None
    return new_id


def update_udhaari(record_id, name, mobile, vehicle, vehicle_no, address,
                    tx_date, due_date, total_amt, paid_amt, due_amt, notes, type, client_id=None,
                    invoice_no="", customer_gstin="", customer_state_code=""):
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''UPDATE udhaari SET
                    name=?, mobile=?, vehicle=?, vehicle_no=?, address=?,
                    tx_date=?, due_date=?, total_amt=?, paid_amt=?, due_amt=?,
                    notes=?, type=?, client_id=?, invoice_no=?, customer_gstin=?, customer_state_code=?
                 WHERE id=?''',
              (name, mobile, vehicle, vehicle_no, address, tx_date,
               due_date, total_amt, paid_amt, due_amt, notes, type, client_id, invoice_no, customer_gstin, customer_state_code, record_id))
    conn.commit()
    conn.close()
    _touch_sync_fields(record_id, "udhaari")
    _queue_sync("udhaari", record_id, "upsert")
    _safe_sync(gsheet_sync.sync_udhaari, get_udhaari()) if _GSHEET_AVAILABLE else None


def delete_udhaari(record_id):
    conn = _get_connection()
    c = conn.cursor()
    c.execute("UPDATE udhaari SET is_deleted=1 WHERE id=?", (record_id,))
    conn.commit()
    conn.close()
    _touch_sync_fields(record_id, "udhaari")
    _queue_sync("udhaari", record_id, "delete")
    _safe_sync(gsheet_sync.sync_udhaari, get_udhaari()) if _GSHEET_AVAILABLE else None


def get_udhaari():
    conn = _get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM udhaari WHERE (is_deleted IS NULL OR is_deleted=0) ORDER BY id DESC")
    data = c.fetchall()
    conn.close()
    return data


def get_by_id(record_id):
    conn = _get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM udhaari WHERE id=?", (record_id,))
    row = c.fetchone()
    conn.close()
    return row


def search_udhaari(query):
    if not query or not query.strip():
        return get_udhaari()
    conn = _get_connection()
    c = conn.cursor()
    like = f"%{query.strip()}%"
    c.execute('''SELECT * FROM udhaari
                 WHERE (name LIKE ? OR vehicle LIKE ? OR vehicle_no LIKE ? OR mobile LIKE ?)
                 AND (is_deleted IS NULL OR is_deleted=0)
                 ORDER BY id DESC''', (like, like, like, like))
    data = c.fetchall()
    conn.close()
    return data


def get_history_by_name(name):
    conn = _get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM udhaari WHERE name=? AND (is_deleted IS NULL OR is_deleted=0) ORDER BY id DESC", (name,))
    data = c.fetchall()
    conn.close()
    return data


# ======================================================================
# CUSTOMER PROFILE / HISTORY
# ======================================================================

def get_all_customer_names():
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT DISTINCT name FROM udhaari
                 WHERE name IS NOT NULL AND name != ''
                 AND (is_deleted IS NULL OR is_deleted=0)
                 ORDER BY name COLLATE NOCASE''')
    data = [row["name"] for row in c.fetchall()]
    conn.close()
    return data


def get_customer_profile(name):
    """एका कस्टमरचा संपूर्ण प्रोफाइल — सगळ्या नोंदी, एकूण दिलेले/घेतलेले/थकीत,
    आणि त्याच्या सगळ्या वेगवेगळ्या गाड्या एकत्र देतो."""
    conn = _get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM udhaari WHERE name=? AND (is_deleted IS NULL OR is_deleted=0) ORDER BY id DESC", (name,))
    records = c.fetchall()
    conn.close()

    if not records:
        return None

    total_given = sum((r["total_amt"] or 0) for r in records if r["type"] == "Given")
    total_taken = sum((r["total_amt"] or 0) for r in records if r["type"] == "Taken")
    total_due = sum((r["due_amt"] or 0) for r in records)
    total_paid = sum((r["paid_amt"] or 0) for r in records)

    vehicles = sorted({
        f"{r['vehicle']} ({r['vehicle_no']})" if r["vehicle_no"] else r["vehicle"]
        for r in records if r["vehicle"]
    })

    latest = records[0]
    return {
        "name": name,
        "mobile": latest["mobile"],
        "address": latest["address"],
        "vehicles": vehicles,
        "total_given": total_given,
        "total_taken": total_taken,
        "total_paid": total_paid,
        "total_due": total_due,
        "record_count": len(records),
        "records": records,
    }


# ======================================================================
# VEHICLE HISTORY
# ======================================================================

def get_all_vehicle_numbers():
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT DISTINCT vehicle_no FROM udhaari
                 WHERE vehicle_no IS NOT NULL AND vehicle_no != ''
                 AND (is_deleted IS NULL OR is_deleted=0)
                 ORDER BY vehicle_no COLLATE NOCASE''')
    data = [row["vehicle_no"] for row in c.fetchall()]
    conn.close()
    return data


def get_vehicle_history(vehicle_no):
    if not vehicle_no or not vehicle_no.strip():
        return []
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT * FROM udhaari WHERE vehicle_no = ? COLLATE NOCASE
                 AND (is_deleted IS NULL OR is_deleted=0)
                 ORDER BY id DESC''', (vehicle_no.strip(),))
    data = c.fetchall()
    conn.close()
    return data


# ======================================================================
# DUE DATE ALERTS
# ======================================================================

def get_due_alerts(within_days=3):
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT * FROM udhaari WHERE due_amt > 0 AND due_date IS NOT NULL
                 AND due_date != '' AND (is_deleted IS NULL OR is_deleted=0)''')
    rows = c.fetchall()
    conn.close()

    today = datetime.now()
    cutoff = today + timedelta(days=within_days)
    upcoming, overdue = [], []

    for row in rows:
        parsed = _parse_date(row["due_date"])
        if parsed is None:
            continue
        if parsed < today:
            overdue.append(row)
        elif parsed <= cutoff:
            upcoming.append(row)

    return {"overdue": overdue, "upcoming": upcoming}


def get_due_soon(within_days=3):
    """due_amt > 0 आणि due_date आजपासून within_days दिवसांत आहे (किंवा आधीच
    उलटून गेली) अशा नोंदी परत देतं — सर्वात तातडीच्या आधी."""
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT * FROM udhaari WHERE due_amt > 0 AND due_date IS NOT NULL
                 AND due_date != '' AND (is_deleted IS NULL OR is_deleted=0)''')
    rows = c.fetchall()
    conn.close()

    today = datetime.now().date()
    cutoff = today + timedelta(days=within_days)
    results = []

    for row in rows:
        parsed = _parse_date(row["due_date"])
        if parsed is None:
            continue
        due_date = parsed.date()
        if due_date <= cutoff:
            days_left = (due_date - today).days
            results.append({
                "row": row, "due_date": due_date,
                "days_left": days_left, "is_overdue": days_left < 0,
            })

    results.sort(key=lambda x: x["days_left"])
    return results


def get_due_customers():
    """ज्यांच्याकडे अजून due_amt > 0 आहे अशा सगळ्या ग्राहकांची यादी
    (नाव, मोबाईल, एकूण थकीत रक्कम) — WhatsApp रिमाइंडरसाठी."""
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT name, mobile, SUM(due_amt) as total_due, type
                 FROM udhaari
                 WHERE due_amt > 0 AND type = 'Given' AND (is_deleted IS NULL OR is_deleted=0)
                 GROUP BY name, mobile
                 ORDER BY total_due DESC''')
    rows = c.fetchall()
    conn.close()
    return rows


# ======================================================================
# SUMMARY / REPORTS
# ======================================================================

def get_summary():
    conn = _get_connection()
    c = conn.cursor()

    c.execute("SELECT COALESCE(SUM(due_amt), 0) FROM udhaari WHERE type='Given' AND (is_deleted IS NULL OR is_deleted=0)")
    total_given_due = c.fetchone()[0]

    c.execute("SELECT COALESCE(SUM(due_amt), 0) FROM udhaari WHERE type='Taken' AND (is_deleted IS NULL OR is_deleted=0)")
    total_taken_due = c.fetchone()[0]

    c.execute("SELECT COALESCE(SUM(total_amt), 0) FROM udhaari WHERE (is_deleted IS NULL OR is_deleted=0)")
    total_business = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM udhaari WHERE (is_deleted IS NULL OR is_deleted=0)")
    total_records = c.fetchone()[0]

    conn.close()
    return {
        "total_given_due": total_given_due,
        "total_taken_due": total_taken_due,
        "total_business": total_business,
        "total_records": total_records,
    }


def get_daily_summary(days=14):
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT tx_date, total_amt, due_amt FROM udhaari
                 WHERE tx_date IS NOT NULL AND tx_date != ''
                 AND (is_deleted IS NULL OR is_deleted=0)''')
    rows = c.fetchall()
    conn.close()

    daily = {}
    for row in rows:
        parsed = _parse_date(row["tx_date"])
        if parsed is None:
            continue
        key = parsed.strftime("%d-%m")
        if key not in daily:
            daily[key] = {"date": parsed, "total": 0.0, "due": 0.0}
        daily[key]["total"] += row["total_amt"] or 0
        daily[key]["due"] += row["due_amt"] or 0

    cutoff = datetime.now() - timedelta(days=days)
    items = [v for v in daily.values() if v["date"] >= cutoff]
    items.sort(key=lambda x: x["date"])

    return {
        "labels": [v["date"].strftime("%d-%m") for v in items],
        "totals": [v["total"] for v in items],
        "dues": [v["due"] for v in items],
    }


def get_monthly_summary(months=6):
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT tx_date, total_amt, due_amt FROM udhaari
                 WHERE tx_date IS NOT NULL AND tx_date != ''
                 AND (is_deleted IS NULL OR is_deleted=0)''')
    rows = c.fetchall()
    conn.close()

    monthly = {}
    for row in rows:
        parsed = _parse_date(row["tx_date"])
        if parsed is None:
            continue
        key = parsed.strftime("%Y-%m")
        if key not in monthly:
            monthly[key] = {"sort_key": key, "label": parsed.strftime("%b %Y"), "total": 0.0, "due": 0.0}
        monthly[key]["total"] += row["total_amt"] or 0
        monthly[key]["due"] += row["due_amt"] or 0

    items = sorted(monthly.values(), key=lambda x: x["sort_key"])[-months:]
    return {
        "labels": [v["label"] for v in items],
        "totals": [v["total"] for v in items],
        "dues": [v["due"] for v in items],
    }


# ======================================================================
# EXCEL EXPORT
# ======================================================================

def export_to_excel(filepath):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    records = get_udhaari()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Udhaari"

    headers = ["ID", "Name", "Mobile", "Vehicle", "Vehicle No", "Address",
               "Trans. Date", "Due Date", "Total Amt", "Paid Amt", "Due Amt",
               "Notes", "Type"]
    sheet.append(headers)

    header_fill = PatternFill("solid", start_color="00FFAA")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in records:
        sheet.append([
            row["id"], row["name"], row["mobile"], row["vehicle"], row["vehicle_no"],
            row["address"], row["tx_date"], row["due_date"], row["total_amt"],
            row["paid_amt"], row["due_amt"], row["notes"], row["type"],
        ])

    widths = [6, 18, 14, 14, 14, 18, 12, 12, 12, 12, 12, 22, 10]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = w

    last_row = sheet.max_row + 2
    sheet.cell(row=last_row, column=8, value="Total:").font = Font(bold=True)
    sheet.cell(row=last_row, column=9, value=f"=SUM(I2:I{sheet.max_row - 1})")
    sheet.cell(row=last_row, column=10, value=f"=SUM(J2:J{sheet.max_row - 1})")
    sheet.cell(row=last_row, column=11, value=f"=SUM(K2:K{sheet.max_row - 1})")

    wb.save(filepath)
    return filepath