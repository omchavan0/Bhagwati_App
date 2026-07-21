"""
============================================================================
DB GST RETURNS — GSTR-1 (B2B/B2C) + GSTR-3B + HSN Summary डेटा तयार करणे
============================================================================
सरकारी पोर्टलवर थेट फाईल करत नाही (ते मॅन्युअली अपलोड/एंट्री करावं लागेल),
पण भरण्यासाठी लागणारे सगळे आकडे इथून अचूक तयार मिळतात. फक्त GST Billing
Screen मधून झालेल्या sales (invoice_no भरलेला) इथे मोजल्या जातात — जुन्या
साध्या Billing/Udhaari नोंदींना invoice-level GST ब्रेकअप उपलब्ध नसतो.
============================================================================
"""
from datetime import datetime

from db_core import _get_connection, _parse_date
from db_inventory import get_part_usage_by_reference, get_gst_summary_report
from db_company import get_company_settings
from gst_utils import determine_tax_mode, split_tax
from db_purchase import get_purchase_bills


def _sales_in_range(date_from=None, date_to=None):
    conn = _get_connection()
    c = conn.cursor()
    c.execute('''SELECT * FROM udhaari
                 WHERE type='Given' AND (is_deleted IS NULL OR is_deleted=0)
                 AND invoice_no IS NOT NULL AND invoice_no != ''
                 ORDER BY id DESC''')
    rows = c.fetchall()
    conn.close()

    if date_from:
        d1 = _parse_date(date_from)
        rows = [r for r in rows if d1 and _parse_date(r["tx_date"]) and _parse_date(r["tx_date"]) >= d1]
    if date_to:
        d2 = _parse_date(date_to)
        rows = [r for r in rows if d2 and _parse_date(r["tx_date"]) and _parse_date(r["tx_date"]) <= d2]
    return rows


def _invoice_tax_breakup(udhaari_row):
    """एका इनव्हॉइसचा Taxable/CGST/SGST/IGST — part_usage स्नॅपशॉट rate/hsn वरून."""
    company = get_company_settings()
    company_state_code = company["state_code"] if company else "27"
    customer_state_code = (udhaari_row["customer_state_code"]
                            if "customer_state_code" in udhaari_row.keys() and udhaari_row["customer_state_code"]
                            else company_state_code)
    tax_mode = determine_tax_mode(company_state_code, customer_state_code)

    items = get_part_usage_by_reference("udhaari", udhaari_row["id"])
    taxable = cgst = sgst = igst = 0.0
    for it in items:
        gst_rate = it["gst_rate"] if "gst_rate" in it.keys() and it["gst_rate"] is not None else 18.0
        split = split_tax(it["net_amount"] or 0, gst_rate, tax_mode, price_includes_gst=True)
        taxable += split["taxable"]; cgst += split["cgst"]
        sgst += split["sgst"]; igst += split["igst"]

    return {"taxable": taxable, "cgst": cgst, "sgst": sgst, "igst": igst,
            "total_tax": cgst + sgst + igst, "tax_mode": tax_mode}


def get_gstr1_b2b(date_from=None, date_to=None):
    """Registered ग्राहकांच्या (GSTIN असलेल्या) इनव्हॉइस-निहाय यादी."""
    result = []
    for r in _sales_in_range(date_from, date_to):
        gstin = r["customer_gstin"] if "customer_gstin" in r.keys() else ""
        if not gstin:
            continue
        tax = _invoice_tax_breakup(r)
        result.append({
            "invoice_no": r["invoice_no"], "invoice_date": r["tx_date"],
            "customer_name": r["name"], "gstin": gstin,
            "taxable_value": tax["taxable"], "cgst": tax["cgst"], "sgst": tax["sgst"],
            "igst": tax["igst"], "total_value": r["total_amt"] or 0,
        })
    return result


def get_gstr1_b2c(date_from=None, date_to=None):
    """Unregistered ग्राहकांच्या इनव्हॉइसेसचं state-निहाय एकत्रित रूप (B2C Small)."""
    groups = {}
    for r in _sales_in_range(date_from, date_to):
        gstin = r["customer_gstin"] if "customer_gstin" in r.keys() else ""
        if gstin:
            continue
        tax = _invoice_tax_breakup(r)
        state_code = (r["customer_state_code"] if "customer_state_code" in r.keys() and r["customer_state_code"] else "27")
        grp = groups.setdefault(state_code, {"taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "count": 0})
        grp["taxable"] += tax["taxable"]; grp["cgst"] += tax["cgst"]
        grp["sgst"] += tax["sgst"]; grp["igst"] += tax["igst"]; grp["count"] += 1

    return [{"state_code": s, **v} for s, v in sorted(groups.items())]


def get_gstr3b_summary(date_from=None, date_to=None):
    """GSTR-3B Table 3.1 (Outward Supplies) — एकूण बेरीज."""
    rows = _sales_in_range(date_from, date_to)
    total_taxable = total_cgst = total_sgst = total_igst = total_value = 0.0
    for r in rows:
        tax = _invoice_tax_breakup(r)
        total_taxable += tax["taxable"]; total_cgst += tax["cgst"]
        total_sgst += tax["sgst"]; total_igst += tax["igst"]
        total_value += r["total_amt"] or 0

    return {
        "total_taxable": total_taxable, "total_cgst": total_cgst,
        "total_sgst": total_sgst, "total_igst": total_igst,
        "total_tax": total_cgst + total_sgst + total_igst,
        "total_invoice_value": total_value, "invoice_count": len(rows),
    }


def get_hsn_summary_for_returns(date_from=None, date_to=None):
    """db_inventory.py मधलाच HSN Summary पुन्हा-वापर (कोड डुप्लिकेट टाळण्यासाठी)."""
    days = None
    if date_from:
        d1 = _parse_date(date_from)
        if d1:
            days = max((datetime.now() - d1).days, 0)
    return get_gst_summary_report(days=days)


def export_gstr1_to_excel(filepath, date_from=None, date_to=None):
    """B2B + B2C दोन्ही एकाच Excel मध्ये, वेगळ्या Sheets मध्ये — CA/Accountant
    कडे पाठवायला तयार फॉरमॅट."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_fill = PatternFill("solid", start_color="00FFAA")

    # ---- B2B Sheet ----
    ws1 = wb.active
    ws1.title = "B2B"
    b2b_headers = ["Invoice No", "Invoice Date", "Customer Name", "GSTIN",
                   "Taxable Value", "CGST", "SGST", "IGST", "Total Value"]
    ws1.append(b2b_headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="000000"); cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in get_gstr1_b2b(date_from, date_to):
        ws1.append([row["invoice_no"], row["invoice_date"], row["customer_name"], row["gstin"],
                    row["taxable_value"], row["cgst"], row["sgst"], row["igst"], row["total_value"]])

    # ---- B2C Sheet ----
    ws2 = wb.create_sheet("B2C")
    b2c_headers = ["State Code", "Taxable Value", "CGST", "SGST", "IGST", "Invoice Count"]
    ws2.append(b2c_headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="000000"); cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in get_gstr1_b2c(date_from, date_to):
        ws2.append([row["state_code"], row["taxable"], row["cgst"], row["sgst"], row["igst"], row["count"]])

    # ---- GSTR-3B Summary Sheet ----
    ws3 = wb.create_sheet("GSTR-3B Summary")
    summary = get_gstr3b_summary(date_from, date_to)
    ws3.append(["Field", "Amount"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="000000"); cell.fill = header_fill
    for label, key in [("Total Taxable Value", "total_taxable"), ("Total CGST", "total_cgst"),
                        ("Total SGST", "total_sgst"), ("Total IGST", "total_igst"),
                        ("Total Tax", "total_tax"), ("Total Invoice Value", "total_invoice_value"),
                        ("Invoice Count", "invoice_count")]:
        ws3.append([label, summary[key]])

    for ws in (ws1, ws2, ws3):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    # export_gstr1_to_excel() मध्ये, ws3 (GSTR-3B Summary) नंतर, wb.save() च्या आधी:

    # ---- Purchase GST Report Sheet ----
    ws4 = wb.create_sheet("Purchase_GST_Report")
    purchase_headers = ["Bill No", "Date", "Supplier Name", "GSTIN",
                         "Taxable Value", "CGST", "SGST", "IGST", "Grand Total"]
    ws4.append(purchase_headers)
    for cell in ws4[1]:
        cell.font = Font(bold=True, color="000000"); cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in get_purchase_gst_register(date_from, date_to):
        ws4.append([row["bill_no"], row["bill_date"], row["supplier_name"], row["supplier_gstin"],
                    row["taxable_value"], row["cgst"], row["sgst"], row["igst"], row["grand_total"]])        

    wb.save(filepath)
    return filepath



def get_purchase_gst_register(date_from=None, date_to=None):
    """Purchase Bills (ITC साठी) — Bill No, Date, Supplier, GSTIN, Taxable,
    CGST/SGST/IGST, Total. Returns (is_return=1) वगळतो — ते वेगळे track होतात."""
    from db_core import _parse_date

    bills = [b for b in get_purchase_bills() if not b["is_return"]]

    if date_from:
        d1 = _parse_date(date_from)
        bills = [b for b in bills if d1 and _parse_date(b["bill_date"]) and _parse_date(b["bill_date"]) >= d1]
    if date_to:
        d2 = _parse_date(date_to)
        bills = [b for b in bills if d2 and _parse_date(b["bill_date"]) and _parse_date(b["bill_date"]) <= d2]

    result = []
    for b in bills:
        result.append({
            "bill_no": b["bill_no"] or "-",
            "bill_date": b["bill_date"] or "-",
            "supplier_name": b["supplier_name"],
            "supplier_gstin": b["supplier_gstin"] if "supplier_gstin" in b.keys() and b["supplier_gstin"] else "-",
            "taxable_value": b["taxable_value"] or 0,
            "cgst": b["cgst"] or 0,
            "sgst": b["sgst"] or 0,
            "igst": b["igst"] or 0,
            "grand_total": b["grand_total"] or 0,
            "itc_eligible": bool(b["itc_eligible"]) if "itc_eligible" in b.keys() and b["itc_eligible"] is not None else True,
        })
    return result